import functools
from pathlib import Path
import openpyxl
import pandas as pd
from loguru import logger
import io
import openpyxl.drawing.image as openpyxl_image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, get_column_letter, range_boundaries
from PIL import Image as PILImage
import matplotlib.pyplot as plt

logger.remove()

VERSION = "2025.09.16.20"


def _supply_temp_context_if_called_outside_cm(func):
    @functools.wraps(func)
    def new_func(*args, **kwargs):
        self: ExcelWriter = args[0]

        # enter temp context if not run within a context manager
        if self.writer is None or self.writer._status != "open":
            # logger.debug("ExcelWriter is not open, opening it now for this method.")
            self.__enter__(mode=("a" if self._has_added_items else None))
            try:
                res = func(*args, **kwargs)
            finally:
                self.__exit__(None, None, None)
        else:
            res = func(*args, **kwargs)

        return res

    return new_func


def _pre_and_post_proc_for_item_addition(func):
    @functools.wraps(func)
    def new_func(*args, **kwargs):
        self: ExcelWriter = args[0]

        # move cursor
        _valid_locations = ["down", "right", None]
        if "location" not in kwargs or kwargs["location"] is None:
            if not self._cursor_ready_to_write:
                self.move_cursor(down=True)
        elif kwargs["location"] not in _valid_locations:
            raise ValueError(
                f"Invalid location {kwargs['location']}. Must be one of {_valid_locations}."
            )
        elif kwargs["location"] == "down":
            self.move_cursor(down=True)
        elif kwargs["location"] == "right":
            self.move_cursor(right=True)
        else:
            raise Exception("Program logic error - this should not happen.")

        # log and run
        logger.debug(
            f"Writing item on {self.sheet_name} at R{self.cur_row}, C{self.cur_col}"
        )

        res = func(*args, **kwargs)

        # set status flags
        self._has_added_items = True
        self._cursor_ready_to_write = False

        return res

    return new_func


class ExcelWriter(object):
    def __init__(
        self,
        file_path: Path,
        mode: str = "w",
        sheet_name: str = "Sheet",
        cur_row: int = 1,
        cur_col: int = 1,
        distance_bw_contents: int = 2,
    ):
        self.file_path = Path(file_path)
        self.mode = mode
        self.sheet_name = sheet_name
        self.cur_row = cur_row
        self.cur_col = cur_col
        self.distance_bw_contents = distance_bw_contents

        # attributes created or used by methods
        self.writer: pd.ExcelWriter = None
        self.ws: Worksheet = None
        self._content_set_sizes = []  # a set is a row of contents
        self._cursor_ready_to_write = True
        self._has_added_items = False

        # hardcoded attributes
        self.start_row = 1
        self.start_col = 1
        self.rows_per_inch = 5  # 6inch, 30 rows
        self.cols_per_inch = 1.6  # 12inch, 19 cols
        self.rows_per_pixel = self.rows_per_inch / 100
        self.cols_per_pixel = self.cols_per_inch / 100

    def __repr__(self):
        return (
            f"ExcelWriter("
            f"file_path='{self.file_path}',"
            f" mode = {self.mode},"
            f" sheet_name={self.sheet_name}),"
            f" cur_row={self.cur_row},"
            f" cur_col={self.cur_col},"
            f" distance_bw_contents={self.distance_bw_contents},"
            f")"
        )

    def __enter__(self, mode: str = None):
        """
        ways to enter:
        1. obj as context manager
        2. manually first entry
        3. manually subsequent entries

        mode in cases:
        1. can be w or a
        2. can be w or a
        3. must be a

        reset_cursor in cases:
        1. either
        2. False
        3. False
        """
        mode = mode or self.mode

        if not self.file_path.exists():
            mode = "w"
            logger.debug(f"File {self.file_path} does not exist, force mode w.")

        logger.debug(f"Enter ExcelWriter: {mode = }")
        self.writer = pd.ExcelWriter(
            self.file_path,
            engine="openpyxl",
            mode=mode,
            if_sheet_exists="overlay" if mode == "a" else None,
        )
        self.writer._status = "open"  # inplant
        self.switch_to_sheet(
            self.sheet_name, create_if_not_exists=True, reset_cursor=False
        )
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        logger.debug("Exit ExcelWriter")
        self.writer.close()
        self.writer._status = "closed"

    @_supply_temp_context_if_called_outside_cm
    def switch_to_sheet(
        self,
        sheet_name: str,
        create_if_not_exists: bool = False,
        reset_cursor: bool = True,
    ):
        if create_if_not_exists and sheet_name not in self.writer.book.sheetnames:
            # logger.debug(f"Sheet {sheet_name} does not exist, creating it.")
            self._make_new_sheet(sheet_name)

        logger.debug(f"Switching to sheet: {sheet_name}")
        assert (
            sheet_name in self.writer.book.sheetnames
        ), f"Sheet {sheet_name} does not exist in the workbook."
        self.ws = self.writer.sheets[sheet_name]
        self.sheet_name = sheet_name
        if reset_cursor:
            self.move_cursor(reset=True)
        return self

    def _make_new_sheet(self, sheet_name: str):
        logger.debug(f"Creating new sheet: {sheet_name}")
        assert (
            sheet_name not in self.writer.book.sheetnames
        ), f"Sheet {sheet_name} already exists in the workbook."
        self.ws = self.writer.book.create_sheet(sheet_name)
        return self

    def move_cursor(
        self,
        reset: bool = False,
        down: bool = False,
        right: bool = False,
        row: int = None,
        col: int = None,
    ):
        if reset:
            self.cur_row = self.start_row
            self.cur_col = self.start_col
            self._content_set_sizes = []  # reset if move to a new sheet
        elif down and right:
            raise ValueError("Cannot move cursor down and right at the same time.")
        elif down:
            if self._content_set_sizes:
                max_height = max(size[0] for size in self._content_set_sizes)
                self.cur_row += max_height + self.distance_bw_contents
                self.cur_col = self.start_col
                self._content_set_sizes = []  # reset after moving down
        elif right:
            if self._content_set_sizes:
                last_width = self._content_set_sizes[-1][1]
                self.cur_col += last_width + self.distance_bw_contents
        elif row is not None or col is not None:
            self.cur_row += row if row is not None else 0
            self.cur_col += col if col is not None else 0
        else:
            raise ValueError("Must specify one of the movements.")

        self._cursor_ready_to_write = True
        return self

    def write(self, item, **kwargs):
        if isinstance(item, pd.DataFrame) or isinstance(
            item, type(pd.DataFrame().style)
        ):
            return self.write_df(item, **kwargs)
        elif isinstance(
            item,
            (plt.Figure, openpyxl_image.Image, PILImage.Image),
        ):
            return self.write_fig(item, **kwargs)
        elif isinstance(item, str):
            return self.write_text(item, **kwargs)
        else:
            raise ValueError(
                f"item must be a DataFrame or Styler or Figure or Image or string, you gave {type(item)}"
            )

    @_pre_and_post_proc_for_item_addition
    @_supply_temp_context_if_called_outside_cm
    def write_df(
        self,
        df: pd.DataFrame,
        title: str = None,
        index=True,
        location: str = None,
        **kwargs,
    ):
        if isinstance(df, pd.DataFrame):
            df_height, df_width = df.shape
        elif isinstance(df, type(pd.DataFrame().style)):
            df_height, df_width = df.data.shape
        else:
            raise ValueError(f"df must be a DataFrame or Styler, you gave {type(df)}")

        self.ws.cell(
            row=self.cur_row, column=self.cur_col, value=title if title else ""
        )
        df.to_excel(
            self.writer,
            sheet_name=self.sheet_name,
            index=index,
            startrow=self.cur_row + 1 - 1,  # 1 for title, -1 for 0-based index
            startcol=self.cur_col,  # table starts at + 1 more col from startcol
            **kwargs,
        )

        self._content_set_sizes.append((df_height + 1, df_width + index + 1))
        return self

    @_pre_and_post_proc_for_item_addition
    @_supply_temp_context_if_called_outside_cm
    def write_fig(
        self,
        fig: plt.Figure | openpyxl_image.Image | PILImage.Image,
        title: str = None,
        location: str = None,
        fig_height: int = None,
        fig_width: int = None,
        scaler: float = None,
        **kwargs,
    ):
        if isinstance(fig, plt.Figure):
            img = _fig_to_img(fig)
        elif isinstance(fig, openpyxl_image.Image):
            img = fig
        elif isinstance(fig, PILImage.Image):
            img = _pil_to_xlimage(fig)
        else:
            raise ValueError(
                f"fig must be a matplotlib Figure or openpyxl Image or PIL Image, you gave {type(fig)}"
            )

        self.ws.cell(
            row=self.cur_row, column=self.cur_col, value=title if title else ""
        )
        if scaler is not None:
            img.width = int(img.width * scaler)
            img.height = int(img.height * scaler)
        self.ws.add_image(
            img,
            row_col_to_cell_ref(
                self.cur_row + 1, self.cur_col + 1
            ),  # 1 col just for titles
        )
        fig_height = (
            int(round(img.height * self.rows_per_pixel, 0))
            if fig_height is None
            else fig_height
        )
        fig_width = (
            int(round(img.width * self.cols_per_pixel, 0))
            if fig_width is None
            else fig_width
        )
        self._content_set_sizes.append((fig_height + 1, fig_width + 1))
        return self

    @_pre_and_post_proc_for_item_addition
    @_supply_temp_context_if_called_outside_cm
    def write_text(self, text: str, title: str = None, location: str = None, **kwargs):
        if not isinstance(text, str):
            raise ValueError(f"text must be a string, you gave {type(text)}")

        self.ws.cell(
            row=self.cur_row, column=self.cur_col, value=title if title else ""
        )
        for i, line in enumerate(text.split("\n")):
            self.ws.cell(row=self.cur_row + 1 + i, column=self.cur_col + 1, value=line)

        text_height = text.count("\n") + 1
        text_width = max(len(line) for line in text.split("\n"))
        self._content_set_sizes.append((text_height + 1, text_width + 1))
        return self

    @_supply_temp_context_if_called_outside_cm
    def auto_fit_column_width(self, max_width: int = 150, fixed_padding: int = 5):
        for column in self.ws.columns:
            max_length = 0
            column = [cell for cell in column]
            # i = 0
            for cell in column:
                # i += 1
                # if i % 50 == 0:
                #     logger.debug(
                #         f"Checked {i} cells in column {column[0].column_letter} so far, current max_length {max_length}"
                #     )
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                if max_length > max_width:
                    max_length = max_width
                    break
            adjusted_width = max_length + fixed_padding
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width
        return self

    def read_df(
        self,
        sheet_name: str = None,
        range_ref: str = None,
        skiprows=None,
        nrows=None,
        usecols=None,
        index_col=None,
    ) -> pd.DataFrame:
        if sheet_name is None:
            sheet_name = self.sheet_name
        return read_excel_df(
            self.file_path,
            sheet_name,
            range_ref,
            skiprows,
            nrows,
            usecols,
            index_col=index_col,
        )

    def read_img(
        self, sheet_name: str = None, anchor_cell: str = None
    ) -> PILImage.Image | None:
        if sheet_name is None:
            sheet_name = self.sheet_name
        return read_excel_img(self.file_path, sheet_name, anchor_cell)

    def read_text(self, sheet_name: str = None, range_ref: str = None) -> str:
        if sheet_name is None:
            sheet_name = self.sheet_name
        if range_ref is None:
            raise ValueError("range_ref must be provided to read text.")
        return read_excel_text(self.file_path, sheet_name, range_ref)


def row_col_to_cell_ref(row, col):
    return get_column_letter(col) + str(row)


def read_excel_df(
    file_path: Path,
    sheet_name: str,
    range_ref: str = None,
    skiprows=None,
    nrows=None,
    usecols=None,
    index_col=None,
) -> pd.DataFrame:
    if not file_path.exists():
        raise FileNotFoundError(f"File {file_path} does not exist.")

    if range_ref is not None:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        skiprows = range(0, min_row - 1)
        nrows = max_row - min_row + 1
        usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
    logger.debug(f"skiprows={skiprows}, nrows={nrows}, usecols={usecols}")
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        skiprows=skiprows,
        nrows=nrows,
        usecols=usecols,
        index_col=index_col,
    )
    return df


def read_excel_img(
    file_path: Path, sheet_name: str, anchor_cell: str = None
) -> PILImage.Image | None:
    if not file_path.exists():
        raise FileNotFoundError(f"File {file_path} does not exist.")
    ws = openpyxl.load_workbook(file_path)[sheet_name]

    _images = getattr(ws, "_images", [])
    logger.debug(f"Found {len(_images)} images in sheet {sheet_name}.")

    for img in _images:
        anchor = getattr(img.anchor, "_from", None)
        if anchor:
            col_letter = chr(65 + anchor.col)
            row_number = anchor.row + 1
            anchored_cell = f"{col_letter}{row_number}"

            if anchor_cell is None or anchored_cell == anchor_cell:
                image_data = img._data()
                image = PILImage.open(io.BytesIO(image_data))
                return image
    return None


def read_excel_text(file_path: Path, sheet_name: str, range_ref: str) -> str:
    if not file_path.exists():
        raise FileNotFoundError(f"File {file_path} does not exist.")
    ws = openpyxl.load_workbook(file_path)[sheet_name]

    min_col, min_row, max_col, max_row = range_boundaries(range_ref)

    texts = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell_value = ws.cell(row=r, column=c).value
            if cell_value is not None:
                texts.append(str(cell_value))
    return "\n".join(texts)


def _fig_to_img(fig: plt.Figure) -> openpyxl_image.Image:
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    img = openpyxl_image.Image(buf)
    return img


def _pil_to_xlimage(pil_img: PILImage.Image) -> openpyxl_image.Image:
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    return openpyxl_image.Image(buf)
