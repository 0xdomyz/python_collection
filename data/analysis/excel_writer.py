import functools
from pathlib import Path
import openpyxl
import pandas as pd
from loguru import logger
from contextlib import contextmanager
import io
from openpyxl.drawing.image import Image as XLImage


VERSION = "2025.08.03.13"


def ref_from_rc(row, col):
    return openpyxl.utils.get_column_letter(col) + str(row)


def _supply_writer_context_for_oneoff_calls(called_flag_name: str):
    def decorate(func):
        @functools.wraps(func)
        def new_func(*args, **kwargs):
            self: ExcelWriter = args[0]
            if self.writer is None or self.writer._status != "open":
                logger.debug("ExcelWriter is not open, opening it now for this method.")
                self.__enter__(
                    mode=(
                        "a" if getattr(self, called_flag_name) else None
                    ),  # a if subsequent call
                    reset_cursor=(
                        False if getattr(self, called_flag_name) else True
                    ),  # no reset if subsequent call
                )
                try:
                    func(*args, **kwargs)
                finally:
                    self.__exit__(None, None, None)
                return self

            return func(*args, **kwargs)

        return new_func

    return decorate


class ExcelWriter(object):
    def __init__(
        self,
        file_path: Path,
        mode: str = "w",
        sheet_name: str = "Sheet",
        cur_row: int = 1,
        cur_col: int = 1,
        distance_bw_contents: int = 2,
    ):
        self.file_path = file_path
        self.writer = None
        self.mode = mode

        # custom attributes
        self.sheet_name = sheet_name
        self.cur_row = cur_row
        self.cur_col = cur_col
        self.distance_bw_contents = distance_bw_contents

        # hardcoded attributes
        self.start_row = 1
        self.start_col = 1
        self.rows_per_inch = 5  # 6inch, 30 rows
        self.cols_per_inch = 1.6  # 12inch, 19 cols

        self._write_df_called = False
        self._write_plot_called = False

    def __repr__(self):
        return (
            f"ExcelWriter(file_path='{self.file_path}', mode = {self.mode},"
            f" sheet_name={self.sheet_name}),"
            f" cur_row={self.cur_row}, cur_col={self.cur_col}, distance_bw_contents={self.distance_bw_contents})"
        )

    # @contextmanager
    # def open(self, mode: str = "a"):
    #     self.__enter__(mode)
    #     try:
    #         yield self
    #     finally:
    #         self.__exit__(None, None, None)

    # def close(self):
    #     if hasattr(self, "writer"):
    #         self.writer.close()
    #     else:
    #         logger.warning("ExcelWriter was not opened, nothing to close.")

    def __enter__(self, mode: str = None, reset_cursor: bool = True):
        """
        ways to enter:
        1. obj as context manager
        2. manually first entry
        3. manually subsequent entries
        """
        mode = mode or self.mode

        if not self.file_path.exists():
            mode = "w"
            logger.debug(f"File {self.file_path} does not exist, force mode w.")

        logger.debug(f"Enter ExcelWriter: {mode = }")
        self.writer = pd.ExcelWriter(
            self.file_path,
            engine="openpyxl",
            mode=mode,
            if_sheet_exists="overlay" if mode == "a" else None,
        )
        self.writer._status = "open"  # inplant
        self.switch_to_sheet(
            self.sheet_name, create_if_not_exists=True, reset_cursor=reset_cursor
        )
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        logger.debug("Exit ExcelWriter")
        self.writer.close()
        self.writer._status = "closed"

    def _make_new_sheet(self, sheet_name: str):
        logger.debug(f"Creating new sheet: {sheet_name}")
        assert (
            sheet_name not in self.writer.book.sheetnames
        ), f"Sheet {sheet_name} already exists in the workbook."
        self.ws = self.writer.book.create_sheet(sheet_name)
        return self

    def switch_to_sheet(
        self,
        sheet_name: str,
        create_if_not_exists: bool = False,
        reset_cursor: bool = True,
    ):
        if create_if_not_exists and sheet_name not in self.writer.book.sheetnames:
            logger.debug(f"Sheet {sheet_name} does not exist, creating it.")
            self._make_new_sheet(sheet_name)

        logger.debug(f"Switching to sheet: {sheet_name}")
        assert (
            sheet_name in self.writer.book.sheetnames
        ), f"Sheet {sheet_name} does not exist in the workbook."
        self.ws = self.writer.sheets[sheet_name]
        self.sheet_name = sheet_name
        if reset_cursor:
            self._reset_cursor()
        return self

    def _reset_cursor(self):
        logger.debug(f"Resetting cursor to {self.start_row}, {self.start_col}")
        self.cur_row = self.start_row
        self.cur_col = self.start_col
        return self

    @_supply_writer_context_for_oneoff_calls("_write_df_called")
    def write_df(self, df: pd.DataFrame, title: str = None, index=True):
        logger.debug(
            f"Writing {title} on {self.sheet_name} at R{self.cur_row}, C{self.cur_col}"
        )

        self.ws.cell(
            row=self.cur_row, column=self.cur_col, value=title if title else ""
        )
        self.cur_row += 1

        df.to_excel(
            self.writer,
            sheet_name=self.sheet_name,
            index=index,
            startrow=self.cur_row - 1,
            startcol=self.cur_col - 1,
        )
        self.cur_row += df.shape[0] + self.distance_bw_contents

        self._write_df_called = True
        return self

    @_supply_writer_context_for_oneoff_calls("_write_plot_called")
    def write_plot(self, fig, title: str = None):
        def fig_to_img(fig):
            buf = io.BytesIO()
            fig.savefig(buf, format="png")
            buf.seek(0)
            img = XLImage(buf)
            return img

        fig_height = int(fig.get_size_inches()[1] * self.rows_per_inch)
        fig_width = int(fig.get_size_inches()[0] * self.cols_per_inch)

        self.ws.cell(
            row=self.cur_row, column=self.cur_col, value=title if title else ""
        )
        self.cur_row += 1

        self.ws.add_image(fig_to_img(fig), ref_from_rc(self.cur_row, self.cur_col))
        self.cur_row += fig_height + self.distance_bw_contents

        self._write_plot_called = True
        return self
