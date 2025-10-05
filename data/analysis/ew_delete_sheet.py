from openpyxl import load_workbook

wb = load_workbook('your_workbook.xlsx')
sheet_name = 'SheetName'  # Replace with the actual sheet name

if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' does not exist.")

if len(wb.sheetnames) == 1:
    raise RuntimeError("Cannot delete the only sheet in the workbook.")

wb.remove(wb[sheet_name])
wb.save('your_workbook.xlsx')