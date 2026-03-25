import unittest
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

MODULE_PATH = Path(__file__).resolve().parent / "pivot.py"
SPEC = spec_from_file_location("pivot_module", MODULE_PATH)
pivot_module = module_from_spec(SPEC)
SPEC.loader.exec_module(pivot_module)

ARTIFACT_DIR = Path(__file__).resolve().parent / "test_artifacts"
ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)


class TestNativeExcelPivotBasic(unittest.TestCase):
    def setUp(self):
        self.df = pd.DataFrame(
            {
                "Region": ["North", "South", "North", "East", "South", "East"],
                "Product": ["A", "A", "B", "B", "B", "A"],
                "Date": [
                    "2025-01-01",
                    "2025-01-15",
                    "2025-02-01",
                    "2025-02-15",
                    "2025-03-01",
                    "2025-03-15",
                ],
                "Sales": [100, 150, 200, 120, 140, 110],
            }
        )

    def test_missing_columns_raises(self):
        with self.assertRaises(ValueError):
            pivot_module.create_native_excel_pivot(
                df=self.df,
                output_file=ARTIFACT_DIR / "basic_missing_columns_dummy.xlsx",
                row_field="Region",
                col_field="MissingColumn",
                value_field="Sales",
            )

    def test_native_pivot_file_and_content(self):
        output_file = ARTIFACT_DIR / "pivot_example_basic.xlsx"

        pivot_module.create_native_excel_pivot(
            df=self.df,
            output_file=output_file,
            row_field="Region",
            col_field="Product",
            value_field="Sales",
        )

        self.assertTrue(output_file.exists())

        wb = load_workbook(output_file, data_only=True)
        self.assertIn("Data", wb.sheetnames)
        self.assertIn("Pivot", wb.sheetnames)

        ws = wb["Pivot"]
        first_col_values = [
            ws.cell(row=row_idx, column=1).value for row_idx in range(1, 20)
        ]
        self.assertIn("Row Labels", first_col_values)
        self.assertIn("East", first_col_values)
        self.assertIn("Grand Total", first_col_values)
        # Verify sum aggregation label is present in pivot
        self.assertIn("Sum of Sales", first_col_values)


if __name__ == "__main__":
    unittest.main(verbosity=2)
