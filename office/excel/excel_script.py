# set up
#################
import numpy as np
import openpyxl
import pandas as pd

here = "office/excel"

# iris data
def iris_df()->pd.DataFrame:
    from sklearn.datasets import load_iris
    _ = load_iris()
    data = pd.DataFrame(
        _.data,
        columns=_.feature_names,
    )
    data["target"] = _.target
    data["target_category"] = data["target"].map(
        {
            0: "setosa",
            1: "versicolor",
            2: "virginica",
        }
    )
    data.columns = [x.replace(" ", "_") for x in data.columns]
    data.columns = [x.replace("(", "") for x in data.columns]
    data.columns = [x.replace(")", "") for x in data.columns]
    return data

iris = iris_df()

# pandas example table
data = pd.DataFrame(
    np.random.randn(10, 4),
    columns=list("ABCD"),
)

# make an excel with the table
data.to_excel(f"{here}/table.xlsx", index=True)

# conditional formatting on existing excel
################################################

wb = openpyxl.load_workbook(f"{here}/table.xlsx")

ws = wb.active

red_text = openpyxl.styles.Font(color="00FF0000")

for row in ws.iter_rows(min_row=2, min_col=2, max_col=5, max_row=11):
    for cell in row:
        if cell.value < 0:
            cell.font = red_text

wb.save(f"{here}/table.xlsx")

# conditional formatting on a dataframe then save to excel
########################################################

# work thru the dataframe, create cells with the color formatting

wb = openpyxl.Workbook()
ws = wb.active

for i in range(data.shape[0]):
    for j in range(data.shape[1]):
        cell = ws.cell(row=i + 2, column=j + 2)
        if data.iloc[i, j] < 0:
            cell.font = red_text
        cell.value = data.iloc[i, j]

wb.save(f"{here}/table2.xlsx")


# use color scales to show the range of values, red to green
############################################################

wb = openpyxl.Workbook()
ws = wb.active

for i in range(data.shape[0]):
    for j in range(data.shape[1]):
        cell = ws.cell(row=i + 2, column=j + 2)
        cell.value = data.iloc[i, j]

ws.conditional_formatting.add(
    f"B2:E{data.shape[0]+1}",
    openpyxl.formatting.rule.ColorScaleRule(
        start_type="min",
        start_color="00FF0000",
        end_type="max",
        end_color="0000FF00",
    ),
)

wb.save(f"{here}/table3.xlsx")


# functionalise, need all numeric data
def save_to_excel_with_color_scale(data: pd.DataFrame, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    # fill in the data
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            cell = ws.cell(row=i + 2, column=j + 2)
            cell.value = data.iloc[i, j]

    # add the color scale
    ws.conditional_formatting.add(
        f"B2:E{data.shape[0]+1}",
        openpyxl.formatting.rule.ColorScaleRule(
            start_type="min",
            start_color="00FF0000",
            end_type="max",
            end_color="0000FF00",
        ),
    )

    wb.save(filename)


data = pd.DataFrame(
    np.random.randn(10, 4),
    columns=list("ABCD"),
)

save_to_excel_with_color_scale(data=data, filename=f"{here}/table4.xlsx")

# add a chart
#############

import openpyxl as xl

data = iris
wb = openpyxl.Workbook()
ws = wb.active
sheetname = "iris"
ws.title = sheetname

# write col name row
for i in range(data.shape[1]):
    ws.cell(row=1, column=i + 2).value = data.columns[i]

# append each row of the dataframe to the sheet, including the index
for i in range(data.shape[0]):
    ws.cell(row=i + 2, column=1).value = data.index[i]
    for j in range(data.shape[1]):
        ws.cell(row=i + 2, column=j + 2).value = data.iloc[i, j]

# todo: make above into a function: dataframe_to_excel(data, filename, sheetname, index=True)

# calculate the table area in A1 notation
def get_table_area(data: pd.DataFrame, include_index=True) -> str:
    if include_index:
        min_col = 1
    else:
        min_col = 2
    max_col = data.shape[1] + 1
    max_row = data.shape[0] + 1
    min_row = 1
    top_left_letter = openpyxl.utils.get_column_letter(min_col)
    down_right_letter = openpyxl.utils.get_column_letter(max_col)
    return f"{top_left_letter}{min_row}:{down_right_letter}{max_row}"

get_table_area(data=data, include_index=True)

def close_open_to_close_close(
        start:int,
        end:int,
)-> tuple:
    """
    close_open_to_close_close(0, 0)
    close_open_to_close_close(0, 1)
    close_open_to_close_close(0, 2)
    close_open_to_close_close(1, 3)
    """
    if start >= end:
        return None
    else:
        return start, end - 1

# column start and end pass 1 number and data, output the range
def get_column_range(data: pd.DataFrame, start: int, end:int, include_index=True) -> str:
    # convert to excel index
    start, end = start + 1, end
    if include_index:
        start += 1
        end += 1
    top_left_letter = openpyxl.utils.get_column_letter(start)
    down_right_letter = openpyxl.utils.get_column_letter(end)
    min_row = 1
    max_row = data.shape[0] + 1
    return f"{top_left_letter}{min_row}:{down_right_letter}{max_row}"

get_column_range(data=data, start=1, end=4)

# add the chart
chart = openpyxl.chart.LineChart()
chart.title = "Iris sepal length vs other measures"
chart.style = 13 # use a preset style: https://openpyxl.readthedocs.io/en/stable/charts/pie.html#pie-chart-styles
chart.x_axis.title = "sepal length (cm)"
chart.y_axis.title = "other measures"
chart.legend.position = "r" # legend on the right

# values are the lines to plot
_ = get_column_range(data=data, start=1, end=4)
values = xl.chart.Reference(ws, range_string=f"{ws.title}!{_}")
_ = get_column_range(data=data, start=0, end=1)
x_values = xl.chart.Reference(ws, range_string=f"{ws.title}!{_}")

chart.add_data(values, titles_from_data=True)
chart.set_categories(x_values)

# add the chart to the sheet
ws.add_chart(chart, "H2")

wb.save(f"{here}/table5.xlsx")
